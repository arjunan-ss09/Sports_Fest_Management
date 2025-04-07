from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db.models import Count, Q
import openpyxl, requests
from django.http import Http404
from django.utils import timezone
from django.contrib.auth import authenticate, login, logout
from .forms import ParticipantRegistrationForm, FeedbackForm
from .forms import MatchForm, ScoreForm
from .models import ParticipantProfile, Event, Team, Match, College, Feedback, TeamJoinRequest
from django.urls import reverse
from django.http import HttpResponseForbidden
import csv
from django.http import HttpResponse
from collections import defaultdict
from io import BytesIO
import openpyxl
from openpyxl.styles import Font
def home(request):
    return render(request, 'management/home.html')

def participant_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            try:
                profile = user.participantprofile
                if profile.is_banned:
                    messages.error(request, "You have been banned from participating in the fest. Kindly contact the organizers.")
                    logout(request)
                    return redirect('participant_login')
            except ParticipantProfile.DoesNotExist:
                pass

            login(request, user)
            return redirect('participant_dashboard')
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, 'account/participant_login.html')

def organizer_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_staff: 
                login(request, user)
                return redirect('organizer_dashboard')
            else:
                messages.error(request, "You are not authorized to access the organizer portal.")
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, 'account/organizer_login.html')

#Participant-related Views
@login_required
def participant_register(request):
    existing_profile = ParticipantProfile.objects.filter(user=request.user).first()
    if existing_profile:
        if existing_profile.is_banned:
            messages.error(request, "You have been banned from participating. Please contact the organizers for more details.")
            return redirect('home')
        else:
            return redirect('participant_dashboard')

    if request.method == "POST":
        form = ParticipantRegistrationForm(request.POST)
        if form.is_valid():
            college_name = request.POST.get("college")
            if college_name.isdigit():
                college = get_object_or_404(College, id=int(college_name))
            else:
                college, created = College.objects.get_or_create(name=college_name)
            participant = form.save(commit=False)
            participant.user = request.user
            participant.college = college
            participant.save()

            messages.success(request, "Registration successful!")
            return redirect("participant_dashboard")
        else:
            print("Form errors:", form.errors)
            messages.error(request, "There were errors in your form. Please fix them below.")
    else:
        form = ParticipantRegistrationForm()

    return render(request, 'management/participant_register.html', {'form': form})

@login_required
def participant_dashboard(request):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "Please complete participant registration first.")
        return redirect('participant_register')
    if profile.is_banned:
        messages.error(request, "You have been banned from accessing the participant portal. Please contact the organizers.")
        return redirect('home')

    my_teams = Team.objects.filter(members=profile).select_related('event')
    upcoming_matches = Match.objects.filter(teams__in=my_teams, finished=False, start_time=timezone.now()).distinct().order_by("start_time")
    
    context = {
        'profile': profile,
        'my_teams': my_teams,
        "upcoming_matches": upcoming_matches,
    }
    return render(request, 'management/participant_dashboard.html', context)

@login_required
def list_events(request):
    try:
        request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "You must be registered as a participant to view events.")
        return redirect('participant_register')

    events = Event.objects.all()
    return render(request, 'management/list_events.html', {'events': events})

@login_required
def join_event(request, event_id):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "Please complete participant registration first.")
        return redirect('participant_register')

    event = get_object_or_404(Event, id=event_id)
    if event.gender != 'Mixed' and event.gender != profile.gender:
        messages.error(request, "Your gender does not match the event requirements.")
        return redirect('list_events')

    team, created = Team.objects.get_or_create(event=event, college=profile.college)
    if not team.members.exists():
        team.captain = profile
        team.save()
        team.members.add(profile)
        messages.success(request, f"You have successfully joined {event.name} as the team captain!")
        return redirect('participant_dashboard')

    if team.is_full():
        messages.error(request, "This team is already full!")
        return redirect('list_events')

    if team.captain:
        TeamJoinRequest.objects.create(team=team, participant=profile)
        messages.success(request, "Your request to join the team has been sent to the captain.")
    else:
        team.captain = profile
        team.save()
        team.members.add(profile)
        messages.success(request, f"You have successfully joined {event.name} as the team captain!")

    return redirect('participant_dashboard')


@login_required
def participant_matches(request):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "Please complete participant registration first.")
        return redirect('participant_register')
    
    my_teams = Team.objects.filter(members=profile)
    matches = Match.objects.filter(
        Q(teams__in=my_teams) | Q(teams__college=profile.college)
    ).distinct().order_by('start_time')
    
    context = {'matches': matches}
    return render(request, 'management/participant_matches.html', context)

@login_required
def view_team(request, team_id):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "Please complete participant registration first.")
        return redirect('participant_register')

    team = get_object_or_404(Team, id=team_id)
    is_captain= (team.captain_id==profile.id)

    pending_requests=team.join_requests.filter(status='pending') if is_captain else None

    if profile not in team.members.all():
        messages.error(request, "You are not a member of this team.")
        return redirect('participant_dashboard')

    return render(request, 'management/team_detail.html', {'team': team, 'is_captain': is_captain, 'pending_requests': pending_requests})

@login_required
def view_join_requests(request):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "You must be a participant to view join requests.")
        return redirect('participant_dashboard')

    team = Team.objects.filter(captain=profile).first()

    if not team:
        messages.error(request, "You are not a team captain.")
        return redirect('participant_dashboard')


    join_requests = TeamJoinRequest.objects.filter(team=team)

    return render(request, 'management/captain_join_requests.html', {'join_requests': join_requests})

@login_required
def handle_join_request(request, request_id):
    join_request = get_object_or_404(TeamJoinRequest, id=request_id)
    profile = request.user.participantprofile

    if join_request.team.captain != profile:
        return HttpResponseForbidden("You are not authorized to process this join request.")

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            join_request.status = 'approved'
            join_request.team.participants.add(join_request.participant)
            messages.success(request, f"{join_request.participant.name} has been added to the team.")
        elif action == 'reject':
            join_request.status = 'rejected'
            messages.info(request, f"{join_request.participant.name}'s join request was rejected.")
        else:
            messages.error(request, "Invalid action.")
        join_request.save()
    return redirect('view_team', team_id=join_request.team.id)

@login_required
def college_participants(request):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "Please complete participant registration first.")
        return redirect('participant_register')

    same_college = ParticipantProfile.objects.filter(college=profile.college)
    from django.utils import timezone
    college_matches = (
        Match.objects.filter(teams__college=profile.college, finished=False, start_time__gte=timezone.now())
        .distinct()
        .order_by("start_time")
    )

    context = {
        'profile': profile,
        'participants': same_college,
        'college_matches': college_matches,
    }
    return render(request, 'management/college_participants.html', context)

@login_required
def submit_feedback(request):
    try:
        profile = request.user.participantprofile
    except ParticipantProfile.DoesNotExist:
        messages.error(request, "Please complete participant registration first.")
        return redirect('participant_register')
    
    if request.method == "POST":
        form = FeedbackForm(request.POST, request.FILES)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.participant = profile
            feedback.save()
            messages.success(request, "Thank you for your feedback!")
            return redirect('participant_dashboard')
        else:
            messages.error(request, "There were errors in your submission. Please correct them.")
    else:
        form = FeedbackForm()
    
    return render(request, 'management/submit_feedback.html', {'form': form})

def calculate_overall_leaderboard():
    from .models import Match
    college_points = defaultdict(float)
    finished_matches = Match.objects.filter(finished=True)
    for match in finished_matches:
        if not match.scores:
            continue
        result = match.scores.get("result")
        teams = list(match.teams.all())
        if len(teams) < 2:
            continue
        if result == "Team1":
            college_points[teams[0].college] += 1.0
            college_points[teams[1].college] += 0.0
        elif result == "Team2":
            college_points[teams[1].college] += 1.0
            college_points[teams[0].college] += 0.0
        elif result == "Draw":
            college_points[teams[0].college] += 0.5
            college_points[teams[1].college] += 0.5
    leaderboard = sorted(college_points.items(), key=lambda x: x[1], reverse=True)
    return leaderboard
def calculate_sport_leaderboard(sport):
    from .models import Match
    college_points = defaultdict(float)
    finished_matches = Match.objects.filter(finished=True, event__sport__iexact=sport)
    for match in finished_matches:
        if not match.scores:
            continue
        result = match.scores.get("result")
        teams = list(match.teams.all())
        if len(teams) < 2:
            continue
        if result == "Team1":
            college_points[teams[0].college] += 1.0
            college_points[teams[1].college] += 0.0
        elif result == "Team2":
            college_points[teams[1].college] += 1.0
            college_points[teams[0].college] += 0.0
        elif result == "Draw":
            college_points[teams[0].college] += 0.5
            college_points[teams[1].college] += 0.5
    leaderboard = sorted(college_points.items(), key=lambda x: x[1], reverse=True)
    return leaderboard

@login_required
def leaderboards(request):
    from .models import Event
    sports = Event.objects.values_list('sport', flat=True).distinct()
    context = {
        'sports': sports,
    }
    return render(request, 'management/leaderboards.html', context)

@login_required
def overall_leaderboard(request):
    leaderboard = calculate_overall_leaderboard()
    return render(request, 'management/overall_leaderboard.html', {'leaderboard': leaderboard})

@login_required
def sport_leaderboard(request, sport):
    leaderboard = calculate_sport_leaderboard(sport)
    return render(request, 'management/sport_leaderboard.html', {'leaderboard': leaderboard, 'sport': sport})

#Organizer-related Views
@staff_member_required
def list_teams(request):
    teams = Team.objects.select_related('event', 'college').all()

    sport_filter = request.GET.get('sport', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    event_filter = request.GET.get('event', '').strip()
    college_filter = request.GET.get('college', '').strip()

    filters = Q()
    if sport_filter:
        filters &= Q(event__sport__icontains=sport_filter)
    if gender_filter:
        filters &= Q(event__gender__iexact=gender_filter)
    if event_filter:
        filters &= Q(event__name__icontains=event_filter)
    if college_filter:
        filters &= Q(college__name__icontains=college_filter)

    teams = teams.filter(filters)

    return render(request, 'management/list_teams.html', {'teams': teams})

@staff_member_required
def team_detail(request, team_id):

    team = get_object_or_404(Team.objects.select_related('event', 'college'), id=team_id)
    members = team.members.all()
    return render(request, 'management/team_detail.html', {'team': team, 'members': members})


@staff_member_required
def organizer_dashboard(request):
    total_participants = ParticipantProfile.objects.filter(is_banned=False).count()
    total_colleges = College.objects.exclude(name="Unknown College").count()
    participants = ParticipantProfile.objects.all()
    colleges = College.objects.exclude(name="Unknown College")
    events = Event.objects.all()
    matches= Match.objects.all()
    search_query = request.GET.get('search', '')
    gender_filter = request.GET.get('gender', '')
    college_filter = request.GET.get('college', '')
    mobile_filter = request.GET.get('mobile', '')
    filters = Q()
    if search_query:
        filters &= Q(name__icontains=search_query)
    if gender_filter:
        filters &= Q(gender__iexact=gender_filter)
    if college_filter:
        filters &= Q(college__id=college_filter)
    if mobile_filter:
        filters &= Q(mobile_number__icontains=mobile_filter)

    participants = participants.filter(filters)
    date_filter = request.GET.get('date', '').strip()
    sport_filter = request.GET.get('sport', '').strip()
    event_filter = request.GET.get('event', '').strip()
    college_filter = request.GET.get('college', '').strip()

    if date_filter:
        matches = matches.filter(start_time__date=date_filter)
    if sport_filter:
        matches = matches.filter(event__sport__icontains=sport_filter)
    if event_filter:
        matches = matches.filter(event__name__icontains=event_filter)
    if college_filter:
        matches = matches.filter(teams__college__name__icontains=college_filter).distinct()

    return render(request, 'management/organizer_dashboard.html', {
        "events": events,
        'total_participants': total_participants,
        'total_colleges': total_colleges,
        'participants': participants,
        'colleges': colleges,
    })

@staff_member_required
def manage_matches(request):
    matches = Match.objects.select_related('event').prefetch_related('teams', 'teams__college', 'teams__members')
    
    date_filter = request.GET.get('date', '').strip()
    sport_filter = request.GET.get('sport', '').strip()
    event_filter = request.GET.get('event', '').strip()
    college_filter = request.GET.get('college', '').strip()

    if date_filter:
        matches = matches.filter(start_time__date=date_filter)
    if sport_filter:
        matches = matches.filter(event__sport__icontains=sport_filter)
    if event_filter:
        matches = matches.filter(event__name__icontains=event_filter)
    if college_filter:
        matches = matches.filter(teams__college__name__icontains=college_filter).distinct()

    all_sports = Event.objects.values_list('sport', flat=True).distinct()
    all_events = Event.objects.all()
    all_colleges = College.objects.all()

    match_form = MatchForm(step=1)
    if request.method == 'POST':
        post_data = request.POST.copy()
        for key in list(post_data.keys()):
            if '__' in key:
                del post_data[key]
        event_id = post_data.get('event')
        match_form = MatchForm(post_data, extra_event_id=event_id)
        if 'choose_teams' in request.POST:
            match_form = MatchForm(post_data, extra_event_id=event_id, step=2)
            if match_form.is_valid():
                pass
            else:
                print("Step 1 errors:", match_form.errors.as_json())
        
        elif 'submit_match' in request.POST:
            match_form = MatchForm(post_data, extra_event_id=event_id, step=2)
            if match_form.is_valid():
                match_form.save()
                messages.success(request, "Match created successfully!")
                return redirect('manage_matches')
            else:
                print("Submission errors:", match_form.errors.as_json())
                messages.error(request, "Please correct the errors in your form.")
        elif 'submit_bulk' in request.POST:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Please upload a valid file for bulk upload.")
            else:
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    sheet = wb.active
                    errors = []
                    row_number = 1
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        row_number += 1
                        event_name, start_time, end_time, team_colleges = row[:4]
                        try:
                            event = Event.objects.get(name=event_name)
                            match = Match.objects.create(
                                event=event,
                                start_time=start_time,
                                end_time=end_time
                            )
                            if team_colleges:
                                for cname in team_colleges.split(','):
                                    cname = cname.strip()
                                    try:
                                        college = College.objects.get(name=cname)
                                        team = Team.objects.get(event=event, college=college)
                                        match.teams.add(team)
                                    except (College.DoesNotExist, Team.DoesNotExist):
                                        errors.append(f"Row {row_number}: No team found for college '{cname}'.")
                            match.save()
                        except Exception as e:
                            errors.append(f"Row {row_number}: Error processing row {row}. Error: {str(e)}")
                    
                    if errors:
                        for error in errors:
                            messages.error(request, error)
                    else:
                        messages.success(request, "Bulk matches uploaded successfully!")
                    return redirect('manage_matches')
                except Exception as e:
                    messages.error(request, f"Error opening file: {str(e)}")
                    
    context = {
        'matches': matches,
        'match_form': match_form,
        'all_sports': all_sports,
        'all_events': all_events,
        'all_colleges': all_colleges,
    }
    return render(request, 'management/manage_matches.html', context)


@staff_member_required
def list_colleges(request):
    colleges = College.objects.all()
    return render(request, 'management/list_colleges.html', {'colleges': colleges})

@staff_member_required
def view_college_participants(request, college_id):
    college = get_object_or_404(College, id=college_id)
    participants = college.participants.all()
    return render(request, 'management/college_participants.html', {'college': college, 'participants': participants})

@staff_member_required
def college_events_detail(request, college_id):
    college = get_object_or_404(College, id=college_id)
    events = Event.objects.filter(team__college=college).distinct()
    return render(request, 'management/college_events_detail.html', {
        'college': college,
        'events': events
    })

@staff_member_required
def college_event_team(request, college_id, event_id):
    try:
        college_id = int(college_id)
        event_id = int(event_id)
    except ValueError:
        raise Http404("Invalid college or event id.")
    team = get_object_or_404(Team, college__pk=college_id, event__pk=event_id)
    return render(request, 'management/team_detail.html', {'team': team})

@staff_member_required
def ajax_load_teams(request):

    event_id = request.GET.get('event_id')
    teams = Team.objects.none()
    if event_id:
        teams = Team.objects.filter(event__id=event_id)
    return render(request, 'management/teams_options.html', {'teams': teams})

@staff_member_required
def view_feedback(request):
    feedbacks = Feedback.objects.select_related('participant').order_by('-created_at')
    return render(request, 'management/view_feedback.html', {'feedbacks': feedbacks})

@staff_member_required
def delete_match(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    if request.method == 'POST':
        match.delete()
        messages.success(request, "Match deleted successfully!")
    else:
        messages.error(request, "Invalid request method.")
    return redirect('manage_matches')

@staff_member_required
def update_score(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    sport = match.event.sport

    if request.method == 'POST':
        form = ScoreForm(sport, request.POST)
        if form.is_valid():
            data = form.cleaned_data
            finished_value = data.pop('finished', False)
            match.scores = data
            match.finished = finished_value
            match.save()
            messages.success(request, "Scores updated successfully!")
            return redirect('manage_matches')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        initial_data = match.scores or {}
        initial_data['finished'] = match.finished
        form = ScoreForm(sport, initial=initial_data)

    return render(request, 'management/update_score.html', {
        'form': form,
        'match': match
    })

@staff_member_required
def remove_player(request, team_id, participant_id):
    team = get_object_or_404(Team, id=team_id)
    participant = get_object_or_404(ParticipantProfile, id=participant_id)
    
    if participant in team.members.all():
        team.members.remove(participant)
        messages.success(request, f"{participant.name} has been removed from the team.")
    else:
        messages.error(request, f"{participant.name} is not a member of this team.")
    
    return redirect('college_event_team', college_id=team.college.id, event_id=team.event.id)

@staff_member_required
def ban_player(request, participant_id):
    profile= get_object_or_404(ParticipantProfile, id=participant_id)
    if not profile.is_banned:
        profile.is_banned = True
        profile.save()
        # Optionally, remove the participant from any teams.
        profile.team_set.clear()
        messages.success(request, f"{profile.name} has been banned.")
    else:
        messages.info(request, f"{profile.name} is already banned.")
    return redirect('organizer_dashboard')
@staff_member_required
def lift_ban(request, participant_id):
    profile = get_object_or_404(ParticipantProfile, id=participant_id)
    if profile.is_banned:
        profile.is_banned = False
        profile.save()
        messages.success(request, f"{profile.name}'s ban has been lifted.")
    else:
        messages.info(request, f"{profile.name} is not banned.")
    return redirect('organizer_dashboard')

@staff_member_required
def download_participants_csv(request):

    participants = ParticipantProfile.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="participants.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Email', 'Gender', 'DOB', 'Mobile', 'College'])
    
    for p in participants:
        writer.writerow([p.name, p.email, p.gender, p.date_of_birth, p.mobile_number, p.college.name])
    
    return response
@staff_member_required
def download_results_csv(request):
    matches = Match.objects.filter(finished=True)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="results.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Event', 'Teams', 'Start Time', 'End Time', 'Scores'])
    
    for m in matches:
        teams_str = "; ".join(str(team) for team in m.teams.all())
        writer.writerow([m.event.name, teams_str, m.start_time, m.end_time, m.scores])
    
    return response


@staff_member_required
def view_results(request):
    from .models import Match
    sport = request.GET.get('sport')
    matches = Match.objects.filter(finished=True).order_by('-start_time')
    if sport:
        matches = matches.filter(event__sport__iexact=sport)
    context = {
        'matches': matches,
        'sport': sport,
    }
    return render(request, 'management/view_results.html', context)


@staff_member_required
def download_leaderboard_excel(request, leaderboard_type, sport=None):

    if leaderboard_type == 'overall':
        leaderboard = calculate_overall_leaderboard()
        filename = 'overall_leaderboard.xlsx'
        title = 'Overall Leaderboard'
    elif leaderboard_type == 'sport' and sport:
        leaderboard = calculate_sport_leaderboard(sport)
        filename = f'{sport}_leaderboard.xlsx'
        title = f'{sport} Leaderboard'
    else:
        from django.shortcuts import redirect
        messages.error(request, "Invalid leaderboard type or missing sport parameter.")
        return redirect('overall_leaderboard')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    headers = ['Rank', 'College', 'Points']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    for idx, (college, points) in enumerate(leaderboard, start=1):
        ws.append([idx, college.name, points])
    
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    response = HttpResponse(
        virtual_workbook.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
